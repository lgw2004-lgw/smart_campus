from utils.base_view import BaseView
from utils.response import success, error, page_response
from utils.pagination import get_page_params
from utils.gen_id import gen_id
from .models import StuStudent, StuStudentFile, StuClass, AcaCourse, AcaScheduling, AcaEnrollment, AcaExam, AcaScore, AcaAttendance, AcaClassroom, AcaPlan, AcaAttendanceSession, AcaExamSignup, AcaWarning, AcaEvaluation, AcaLeaveApply, AcaLeaveApproval
import json

def _get_jwc_college(request):
    """若当前登录为教务处，返回其所属学院dept_id，否则None（管理员看全部）"""
    try:
        from system_app.models import SysUser, SysDept, SysRole, SysRoleUser
        user_id=None
        user_type=None
        # 从中间件注入的user_info取
        info=getattr(request, 'user_info', None)
        if info and isinstance(info, dict):
            user_id=info.get('userId') or info.get('user_id')
            user_type=str(info.get('user_type') or info.get('userType') or '')
        # 尝试从token头解析
        if not user_id:
            token=request.META.get('HTTP_TOKEN') or request.headers.get('token') or ''
            if token:
                try:
                    from utils.auth import decode_token
                    payload=decode_token(token)
                    user_id=payload.get('userId')
                    user_type=str(payload.get('user_type') or payload.get('userType') or '')
                except:
                    pass
        if not user_id:
            return None
        # 判断是否为教务处：user_type 6 或角色含教务
        is_jwc=False
        if str(user_type)=='6':
            is_jwc=True
        else:
            try:
                rids=list(SysRoleUser.objects.filter(user_id=user_id).values_list('role_id', flat=True))
                if rids:
                    for r in SysRole.objects.filter(role_id__in=rids):
                        if '教务' in r.role_name:
                            is_jwc=True
                            break
            except:
                pass
        if not is_jwc:
            return None
        # 取用户所属院系
        try:
            u=SysUser.objects.get(user_id=user_id)
            if not u.dept_id:
                return None
            d=SysDept.objects.get(dept_id=u.dept_id)
            college_id=d.dept_id if d.parent_id==0 else d.parent_id
            return college_id
        except:
            return None
    except:
        return None

def _allowed_dept_ids(college_id):
    if not college_id:
        return None
    try:
        from system_app.models import SysDept
        ids=[college_id]
        for d in SysDept.objects.filter(parent_id=college_id):
            ids.append(d.dept_id)
        return ids
    except:
        return [college_id]

def _current_user_ctx(request):
    """返回当前用户上下文：user_id/user_type/college/is_admin/is_jwc/is_teacher"""
    info = getattr(request, 'user_info', None)
    user_id = None; user_type = ''
    if info and isinstance(info, dict):
        user_id = info.get('userId') or info.get('user_id')
        user_type = str(info.get('user_type') or info.get('userType') or '')
    if not user_id:
        token = request.META.get('HTTP_TOKEN') or request.headers.get('token') or ''
        if token:
            try:
                from utils.auth import decode_token
                payload = decode_token(token)
                user_id = payload.get('userId')
                user_type = str(payload.get('user_type') or payload.get('userType') or '')
            except:
                pass
    ctx = {'user_id': user_id, 'user_type': user_type, 'college': None,
           'is_admin': user_type in ('1', '0'), 'is_jwc': False, 'is_teacher': str(user_type) == '7'}
    if str(user_type) == '6':
        ctx['is_jwc'] = True
        ctx['college'] = _get_jwc_college(request)
    if str(user_type) == '7':
        # 教师所属学院
        ctx['college'] = _user_college(request)
    return ctx

def _user_college(request):
    """返回当前登录用户所属学院dept_id（含教师/教务/其他），否则None"""
    try:
        from system_app.models import SysUser, SysDept
        info = getattr(request, 'user_info', None)
        user_id = None
        if info and isinstance(info, dict):
            user_id = info.get('userId') or info.get('user_id')
        if not user_id:
            token = request.META.get('HTTP_TOKEN') or request.headers.get('token') or ''
            if token:
                try:
                    from utils.auth import decode_token
                    user_id = decode_token(token).get('userId')
                except:
                    pass
        if not user_id:
            return None
        u = SysUser.objects.filter(user_id=user_id).first()
        if not u or not u.dept_id:
            return None
        d = SysDept.objects.filter(dept_id=u.dept_id).first()
        if not d:
            return None
        return d.dept_id if d.parent_id == 0 else d.parent_id
    except:
        return None

# ---- 学生建档 ----
class StudentQueryByIdCardView(BaseView):
    def get(self, request):
        id_card = request.GET.get('idCard') or request.GET.get('id_card')
        if not id_card:
            return error("idCard 不能为空", code=400)
        try:
            stu = StuStudent.objects.get(id_card=id_card)
            return success({k: getattr(stu, k) for k in ['student_id','name','sex','id_card','phone','dept_id','class_id','enroll_year','avatar']})
        except StuStudent.DoesNotExist:
            return error("未找到学生", code=404)

class StudentQueryByIdView(BaseView):
    def get(self, request, student_id):
        try:
            stu = StuStudent.objects.get(student_id=student_id)
            data={k: getattr(stu, k) for k in ['student_id','name','sex','id_card','phone','dept_id','class_id','enroll_year','avatar','is_final','create_time']}
            # 关联档案
            try:
                from .models import StuStudentFile
                f=StuStudentFile.objects.get(student_id=student_id)
                data.update({k: getattr(f,k) for k in ['family_info','health_info','award_punish','remark','emergency_contact','emergency_phone']})
            except: pass
            # 学院/专业名
            try:
                from system_app.models import SysDept
                if stu.dept_id:
                    d=SysDept.objects.filter(dept_id=stu.dept_id).first()
                    if d:
                        data['dept_name']=d.dept_name
                        if d.parent_id and d.parent_id!=0:
                            p=SysDept.objects.filter(dept_id=d.parent_id).first()
                            if p: data['college_name']=p.dept_name
                            data['major_name']=d.dept_name
                        else:
                            data['college_name']=d.dept_name
                if stu.class_id:
                    from .models import StuClass
                    c=StuClass.objects.filter(class_id=stu.class_id).first()
                    if c: data['class_name']=c.class_name
            except: pass
            return success(data)
        except StuStudent.DoesNotExist:
            return error("未找到学生", code=404)

class StudentAddView(BaseView):
    def post(self, request):
        body = self.parse_body(request)
        sid = body.get('studentId') or body.get('student_id')
        # 编辑分支：学号已存在则更新
        if sid and StuStudent.objects.filter(student_id=sid).exists():
            # 身份证唯一校验（排除自身）
            id_card = body.get('idCard') or body.get('id_card')
            if id_card and StuStudent.objects.filter(id_card=id_card).exclude(student_id=sid).exists():
                return error("身份证已存在", code=400)
            StuStudent.objects.filter(student_id=sid).update(
                name=body.get('name') or body.get('studentName') or StuStudent.objects.get(student_id=sid).name,
                sex=body.get('sex') if body.get('sex') is not None else StuStudent.objects.get(student_id=sid).sex,
                id_card=id_card if id_card is not None else StuStudent.objects.get(student_id=sid).id_card,
                phone=body.get('phone') if body.get('phone') is not None else StuStudent.objects.get(student_id=sid).phone,
                dept_id=body.get('deptId') if body.get('deptId') is not None else StuStudent.objects.get(student_id=sid).dept_id,
                class_id=body.get('classId') if body.get('classId') is not None else StuStudent.objects.get(student_id=sid).class_id,
                enroll_year=body.get('enrollYear') if body.get('enrollYear') is not None else StuStudent.objects.get(student_id=sid).enroll_year,
                avatar=body.get('avatar') if body.get('avatar') is not None else StuStudent.objects.get(student_id=sid).avatar,
            )
            return success({"studentId": sid})
        # 新增分支
        if not sid:
            sid = gen_id('STU')
        if StuStudent.objects.filter(student_id=sid).exists():
            return error("学号已存在", code=400)
        if body.get('idCard') and StuStudent.objects.filter(id_card=body['idCard']).exists():
            return error("身份证已存在", code=400)
        stu = StuStudent.objects.create(
            student_id=sid,
            name=body.get('name') or body.get('studentName'),
            sex=body.get('sex','0'),
            id_card=body.get('idCard'),
            phone=body.get('phone'),
            dept_id=body.get('deptId'),
            class_id=body.get('classId'),
            enroll_year=body.get('enrollYear'),
            avatar=body.get('avatar'),
        )
        return success({"studentId": stu.student_id})

class StudentQueryByPageView(BaseView):
    def post(self, request):
        page_no, page_size, data = get_page_params(request)
        qs = StuStudent.objects.all()
        # 教务处仅看本学院
        college=_get_jwc_college(request)
        if college:
            allowed=_allowed_dept_ids(college)
            if allowed is not None:
                qs=qs.filter(dept_id__in=allowed)
        if data.get('name'):
            qs = qs.filter(name__icontains=data['name'])
        if data.get('collegeId'):
            try:
                from system_app.models import SysDept
                mids=list(SysDept.objects.filter(parent_id=data['collegeId']).values_list('dept_id', flat=True))
                qs = qs.filter(dept_id__in=mids) if mids else qs.none()
            except:
                pass
        if data.get('deptId') or data.get('majorId'):
            mid=data.get('deptId') or data.get('majorId')
            qs = qs.filter(dept_id=mid)
        if data.get('classId'):
            qs = qs.filter(class_id=data['classId'])
        total = qs.count()
        lst = list(qs.order_by('-create_time')[(page_no-1)*page_size: page_no*page_size].values())
        return page_response(lst, total, page_no, page_size)

class StudentFileAddView(BaseView):
    def put(self, request):
        body = self.parse_body(request)
        sid = body.get('studentId') or body.get('student_id')
        if not sid:
            return error("studentId 不能为空", code=400)
        defaults = {
            'family_info': body.get('familyInfo') or body.get('family_info'),
            'health_info': body.get('healthInfo') or body.get('health_info'),
            'award_punish': body.get('awardPunish') or body.get('award_punish'),
            'remark': body.get('remark'),
            'emergency_contact': body.get('emergencyContact') or body.get('emergency_contact'),
            'emergency_phone': body.get('emergencyPhone') or body.get('emergency_phone'),
        }
        obj, created = StuStudentFile.objects.update_or_create(student_id=sid, defaults=defaults)
        return success({"studentId": sid})

    def post(self, request):
        return self.put(request)

class StudentFileQueryView(BaseView):
    def get(self, request, stu_id):
        try:
            f = StuStudentFile.objects.get(student_id=stu_id)
            return success({k: getattr(f,k) for k in ['student_id','family_info','health_info','award_punish','remark','emergency_contact','emergency_phone','update_time']})
        except StuStudentFile.DoesNotExist:
            return success({})

# ---- 课程 ----
class CourseQueryByPageView(BaseView):
    def post(self, request):
        page_no, page_size, data = get_page_params(request)
        qs = AcaCourse.objects.all()
        college=_get_jwc_college(request)
        if college:
            allowed=_allowed_dept_ids(college)
            if allowed is not None:
                qs=qs.filter(dept_id__in=allowed)
        if data.get('courseName'):
            qs = qs.filter(course_name__icontains=data['courseName'])
        if data.get('courseType'):
            qs = qs.filter(course_type=data['courseType'])
        if data.get('status'):
            qs = qs.filter(status=data['status'])
        total = qs.count()
        lst = list(qs.order_by('-create_time')[(page_no-1)*page_size: page_no*page_size].values())
        return page_response(lst, total, page_no, page_size)

class CourseSaveView(BaseView):
    def post(self, request):
        body = self.parse_body(request)
        cid = body.get('courseId') or gen_id('COUR')
        course_type = body.get('courseType')
        if AcaCourse.objects.filter(course_id=cid).exists():
            AcaCourse.objects.filter(course_id=cid).update(course_name=body.get('courseName'), credit=body.get('credit'), hours=body.get('hours'), dept_id=body.get('deptId'), course_type=course_type, status=body.get('status','0'))
            # 同步 course_code 若传
            if body.get('courseCode'):
                AcaCourse.objects.filter(course_id=cid).update(course_code=body.get('courseCode'))
            return success({"courseId": cid})
        try:
            c = AcaCourse.objects.create(course_id=cid, course_name=body['courseName'], course_code=body.get('courseCode') or cid, credit=body.get('credit',3), hours=body.get('hours',48), dept_id=body.get('deptId'), course_type=course_type, status=body.get('status','0'))
        except Exception as e:
            return error(str(e))
        return success({"courseId": c.course_id})

class CourseDeleteView(BaseView):
    def post(self, request, course_id=None):
        cid = course_id or self.parse_body(request).get('courseId') or self.parse_body(request).get('course_id')
        if not cid:
            return error("courseId 不能为空", code=400)
        try:
            c = AcaCourse.objects.get(course_id=cid)
        except AcaCourse.DoesNotExist:
            return error("课程不存在", code=404)
        if AcaScore.objects.filter(course_id=cid).exists():
            return error("该课程存在成绩记录，无法删除，请先删除成绩或标记停用", code=400)
        if AcaEnrollment.objects.filter(course_id=cid, status__in=['0','1']).exists():
            return error("该课程存在有效选课（待缴费/已选），无法删除", code=400)
        if AcaScheduling.objects.filter(course_id=cid).exists():
            # 级联删除排课
            AcaScheduling.objects.filter(course_id=cid).delete()
        c.delete()
        return success()

class CourseSelectableView(BaseView):
    """GET /course/querySelectable?studentId=&semester="""
    def get(self, request):
        student_id = request.GET.get('studentId') or request.GET.get('student_id')
        # 过滤已选、容量、时间冲突简化：仅过滤已选
        enrolled_course_ids = list(AcaEnrollment.objects.filter(student_id=student_id, status__in=['0','1']).values_list('course_id', flat=True))
        qs = AcaCourse.objects.filter(status='0')
        if enrolled_course_ids:
            qs = qs.exclude(course_id__in=enrolled_course_ids)
        lst = list(qs.values())
        return success(lst)

# ---- 排课 ----
class SchedulingSelectView(BaseView):
    def post(self, request):
        body = self.parse_body(request)
        qs = AcaScheduling.objects.all()
        college=_get_jwc_college(request)
        if college:
            allowed=_allowed_dept_ids(college)
            if allowed is not None:
                allowed_courses=list(AcaCourse.objects.filter(dept_id__in=allowed).values_list('course_id', flat=True))
                qs=qs.filter(course_id__in=allowed_courses)
        if body.get('courseId'):
            qs = qs.filter(course_id=body['courseId'])
        if body.get('teacherId'):
            qs = qs.filter(teacher_id=body['teacherId'])
        if body.get('semester'):
            qs = qs.filter(semester=body['semester'])
        if body.get('majorId'):
            qs = qs.filter(major_id=body['majorId'])
        if body.get('weekday'):
            qs = qs.filter(weekday=body['weekday'])
        if body.get('isPublished'):
            qs = qs.filter(is_published=body['isPublished'])
        lst = list(qs.order_by('weekday','section_type').values())
        return success(lst)

class SchedulingAddView(BaseView):
    def post(self, request):
        body = self.parse_body(request)
        def gv(camel, snake): return body.get(camel) if body.get(camel) is not None else body.get(snake)
        teacher_id = gv('teacherId','teacher_id')
        section_type = gv('sectionType','section_type') or '1'
        course_id = gv('courseId','course_id')
        classroom_id = gv('classroomId','classroom_id')
        semester = gv('semester','semester')
        weekday = gv('weekday','weekday')
        start_week = gv('startWeek','start_week')
        end_week = gv('endWeek','end_week')
        major_id = gv('majorId','major_id')
        capacity = gv('capacity','capacity')
        scheduling_day = gv('schedulingDay','scheduling_day')
        if not course_id or not teacher_id or not semester:
            return error("courseId/teacherId/semester必填", code=400)
        sid = gv('id','id')
        if sid:
            AcaScheduling.objects.filter(id=sid).update(course_id=course_id, teacher_id=teacher_id, classroom_id=classroom_id, semester=semester, weekday=weekday, start_week=start_week, end_week=end_week, major_id=major_id, capacity=capacity, section_type=section_type, scheduling_day=scheduling_day)
            return success({"id": int(sid)})
        obj = AcaScheduling.objects.create(course_id=course_id, teacher_id=teacher_id, classroom_id=classroom_id, semester=semester, weekday=weekday, start_week=start_week, end_week=end_week, major_id=major_id, capacity=capacity, section_type=section_type, scheduling_day=scheduling_day)
        return success({"id": obj.id})

class SchedulingUpdateView(BaseView):
    def post(self, request):
        body = self.parse_body(request)
        sid = body.get('id')
        if not sid:
            return error("id 不能为空", code=400)
        # 兼容 camel/snake 双命名
        mapping = {'courseId':'course_id','course_id':'course_id','teacherId':'teacher_id','teacher_id':'teacher_id','classroomId':'classroom_id','classroom_id':'classroom_id','schedulingDay':'scheduling_day','scheduling_day':'scheduling_day','sectionType':'section_type','section_type':'section_type','schedulingType':'scheduling_type','scheduling_type':'scheduling_type'}
        upd={}
        for k,v in body.items():
            if k in mapping:
                upd[mapping[k]]=v
        if upd:
            AcaScheduling.objects.filter(id=sid).update(**upd)
        return success()

# ---- 选课 ----
class EnrollmentAddView(BaseView):
    def post(self, request):
        body = self.parse_body(request)
        student_id = body.get('studentId') or body.get('student_id')
        schedule_id = body.get('scheduleId') or body.get('schedule_id')
        if not student_id or not schedule_id:
            return error("studentId与scheduleId必填", code=400)
        try:
            sch = AcaScheduling.objects.get(id=schedule_id)
        except AcaScheduling.DoesNotExist:
            return error("排课记录不存在", code=404)
        course_id = sch.course_id
        if sch.is_published != '1':
            return error("该课程尚未发布，无法选课", code=400)
        try:
            stu = StuStudent.objects.get(student_id=student_id)
            if stu.is_final == '1':
                return error("课程已完成，无法重复修读（已归档）", code=400)
            stu_major = None
            if stu.class_id:
                c = StuClass.objects.filter(class_id=stu.class_id).first()
                if c: stu_major = c.dept_id
        except StuStudent.DoesNotExist:
            return error("学生不存在", code=404)
        # 专业课仅本专业可选；公共/通识课全校可修
        try:
            crs = AcaCourse.objects.get(course_id=course_id)
            if crs.status == '1':
                return error("课程已停用，无法选课", code=400)
            if crs.course_type not in ('public_basic','general') and sch.major_id and stu_major and sch.major_id != stu_major:
                return error("该课程仅面向开课专业学生", code=400)
        except AcaCourse.DoesNotExist:
            return error("课程不存在", code=404)
        if AcaEnrollment.objects.filter(student_id=student_id, course_id=course_id, status__in=['0','1']).exists():
            return error("已选该课程", code=400)
        # 容量校验
        if sch.capacity:
            cnt = AcaEnrollment.objects.filter(schedule_id=schedule_id, status__in=['0','1']).count()
            if cnt >= sch.capacity:
                return error("该教学班已满员", code=400)
        # 总学费校验：需先缴纳当前学期总学费（TUITION 订单已付）
        try:
            from finance_app.models import FeeOrder, FeeTuitionConfig, FeeOrderItem
            cfg = FeeTuitionConfig.objects.order_by('-create_time').first()
            if cfg:
                sem = cfg.semester
                has_tuition = FeeOrder.objects.filter(student_id=student_id, semester=sem, order_type='TUITION', order_status='3').exists()
                if not has_tuition:
                    # 未缴总学费禁止选课（重修也需先缴总学费）
                    return error(f"请先缴纳 {sem} 总学费", code=400)
        except Exception:
            pass
        # 是否重修：已有成绩即视为重修（需另缴重修费）
        is_retake = AcaScore.objects.filter(student_id=student_id, course_id=course_id).exists()
        enroll_id = gen_id('ENR')
        AcaEnrollment.objects.create(enroll_id=enroll_id, student_id=student_id, course_id=course_id, schedule_id=schedule_id, status='0')
        if is_retake:
            # 自动计算重修费并生成 RETAKE 订单（待付）
            try:
                from finance_app.models import FeeOrder, FeeOrderItem
                price = float(crs.credit or 3) * 100
                order_id = gen_id('ORD')
                sem = cfg.semester if 'cfg' in locals() and cfg else '2024-2025-1'
                FeeOrder.objects.create(order_id=order_id, student_id=student_id, order_amount=price, order_status='0', order_type='RETAKE', semester=sem, ch_id=enroll_id, detail=f"重修 {crs.course_name}")
                FeeOrderItem.objects.create(item_id=gen_id('ITEM'), order_id=order_id, ref_id=enroll_id, item_name=f"重修-{crs.course_name}", item_price=price, item_num=1, item_amount=price)
                return success({"enrollId": enroll_id, "isRetake": True, "retakeFee": price, "retakeOrderId": order_id})
            except Exception as e:
                return success({"enrollId": enroll_id, "isRetake": True, "retakeFeeError": str(e)})
        return success({"enrollId": enroll_id, "isRetake": False})

class EnrollmentCancelView(BaseView):
    def post(self, request, enroll_id):
        try:
            en = AcaEnrollment.objects.get(enroll_id=enroll_id)
        except AcaEnrollment.DoesNotExist:
            return error("选课记录不存在", code=404)
        if en.status != '0':
            return error("仅待缴费可退选", code=400)
        en.status = '2'
        en.save(update_fields=['status'])
        return success()

class EnrollmentQueryByPageView(BaseView):
    def post(self, request):
        page_no, page_size, data = get_page_params(request)
        qs = AcaEnrollment.objects.all()
        college=_get_jwc_college(request)
        if college:
            allowed=_allowed_dept_ids(college)
            if allowed is not None:
                allowed_courses=list(AcaCourse.objects.filter(dept_id__in=allowed).values_list('course_id', flat=True))
                qs=qs.filter(course_id__in=allowed_courses)
        if data.get('studentId'):
            qs = qs.filter(student_id=data['studentId'])
        if data.get('courseId'):
            qs = qs.filter(course_id=data['courseId'])
        if data.get('status') not in (None, '') :
            qs = qs.filter(status=data['status'])
        total = qs.count()
        lst = list(qs.order_by('-create_time')[(page_no-1)*page_size: page_no*page_size].values())
        return page_response(lst, total, page_no, page_size)

class EnrollmentWorkNumView(BaseView):
    """POST /enrollment/queryWorkNum 按课程聚合选课人数，联表返回课程名（剔除无课程的脏数据）"""
    def post(self, request):
        from django.db.models import Count
        valid_ids = AcaCourse.objects.values_list('course_id', flat=True)
        college=_get_jwc_college(request)
        if college:
            allowed=_allowed_dept_ids(college)
            if allowed is not None:
                valid_ids = AcaCourse.objects.filter(dept_id__in=allowed).values_list('course_id', flat=True)
        rows = (AcaEnrollment.objects.filter(status__in=['0','1'], course_id__in=valid_ids)
                .values('course_id').annotate(cnt=Count('enroll_id')).order_by('course_id'))
        # name_map按学院过滤
        if college and allowed is not None:
            name_map = {c.course_id: c.course_name for c in AcaCourse.objects.filter(dept_id__in=allowed)}
        else:
            name_map = {c.course_id: c.course_name for c in AcaCourse.objects.all()}
        out = [{'courseId': r['course_id'], 'courseName': name_map.get(r['course_id']) or '未知课程', 'cnt': r['cnt']} for r in rows]
        return success(out)

# ---- 考试 ----
def _exam_college_scope_qs(qs, request):
    college=_get_jwc_college(request)
    if college:
        allowed=_allowed_dept_ids(college)
        if allowed is not None:
            allowed_courses=list(AcaCourse.objects.filter(dept_id__in=allowed).values_list('course_id', flat=True))
            qs=qs.filter(course_id__in=allowed_courses)
    return qs

class ExamQueryByPageView(BaseView):
    def post(self, request):
        page_no, page_size, data = get_page_params(request)
        qs = AcaExam.objects.all()
        qs = _exam_college_scope_qs(qs, request)
        if data.get('courseId'):
            qs = qs.filter(course_id=data['courseId'])
        if data.get('semester'):
            qs = qs.filter(semester=data['semester'])
        if data.get('examType'):
            qs = qs.filter(exam_type=str(data['examType']))
        if data.get('status'):
            qs = qs.filter(status=data['status'])
        if data.get('roomNo'):
            qs = qs.filter(room_no__icontains=data['roomNo'])
        total = qs.count()
        lst = list(qs.order_by('-exam_date','start_time')[(page_no-1)*page_size: page_no*page_size].values())
        return page_response(lst, total, page_no, page_size)

class ExamSaveView(BaseView):
    def post(self, request):
        body = self.parse_body(request)
        course_id = body.get('courseId')
        if not course_id:
            return error("courseId 必填", code=400)
        # 课程存在性 + 学院
        course = AcaCourse.objects.filter(course_id=course_id).first()
        if not course:
            return error("课程不存在", code=404)
        ctx = _current_user_ctx(request)
        # 教务只能为本院课程排考
        if ctx['is_jwc'] and ctx['college']:
            course_college = course.dept_id if (course.dept_id and course.dept_id < 100) else None
            # dept_id 落在学院级（<1000 视为学院；专业级为 1000+）
            if course.dept_id and course.dept_id >= 1000:
                from system_app.models import SysDept
                d = SysDept.objects.filter(dept_id=course.dept_id).first()
                course_college = d.parent_id if d and d.parent_id else course.dept_id
            if course_college != ctx['college']:
                return error("无权限：仅能为本院课程安排考试", code=403)
        eid = body.get('examId') or gen_id('EXAM')
        exam_date = body.get('examDate')
        start_time = body.get('startTime')
        end_time = body.get('endTime')
        classroom_id = body.get('classroomId')
        room_no = body.get('roomNo')
        if classroom_id and not room_no:
            cr = AcaClassroom.objects.filter(classroom_id=classroom_id).first()
            if cr: room_no = cr.room_no
        # 计算 exam_time
        exam_time = None
        if exam_date and start_time:
            try:
                from datetime import datetime
                exam_time = datetime.strptime(f"{exam_date} {start_time}", "%Y-%m-%d %H:%M:%S" if len(str(start_time))>8 else "%Y-%m-%d %H:%M")
            except:
                try:
                    from datetime import datetime
                    exam_time = datetime.strptime(f"{exam_date} {start_time}", "%Y-%m-%d %H:%M")
                except:
                    exam_time = None
        college_id = course.dept_id
        if college_id and college_id >= 1000:
            from system_app.models import SysDept
            d = SysDept.objects.filter(dept_id=college_id).first()
            if d and d.parent_id: college_id = d.parent_id
        defaults = dict(
            course_id=course_id,
            exam_name=body.get('examName') or (course.course_name + ' 考试'),
            semester=body.get('semester'),
            exam_type=str(body.get('examType') or '1'),
            exam_date=exam_date,
            start_time=start_time,
            end_time=end_time,
            classroom_id=classroom_id,
            room_no=room_no,
            college_id=college_id,
            exam_time=exam_time,
            status=body.get('status') or '0',
        )
        if AcaExam.objects.filter(exam_id=eid).exists():
            AcaExam.objects.filter(exam_id=eid).update(**defaults)
        else:
            defaults['exam_id'] = eid
            defaults['create_by'] = ctx.get('user_id')
            AcaExam.objects.create(**defaults)
        return success({"examId": eid})

class ExamPublishView(BaseView):
    def post(self, request):
        body = self.parse_body(request)
        eid = body.get('examId')
        status = body.get('status') or '1'
        exam = AcaExam.objects.filter(exam_id=eid).first()
        if not exam:
            return error("考试不存在", code=404)
        ctx = _current_user_ctx(request)
        if ctx['is_jwc'] and ctx['college'] and exam.college_id != ctx['college']:
            return error("无权限：仅能发布本院考试", code=403)
        AcaExam.objects.filter(exam_id=eid).update(status=status)
        return success({"examId": eid, "status": status})

class ExamDeleteView(BaseView):
    def post(self, request, exam_id=None):
        eid = exam_id or self.parse_body(request).get('examId') or self.parse_body(request).get('exam_id')
        if not eid:
            return error("examId 不能为空", code=400)
        exam = AcaExam.objects.filter(exam_id=eid).first()
        if not exam:
            return error("考试不存在", code=404)
        ctx = _current_user_ctx(request)
        if ctx['is_jwc'] and ctx['college'] and exam.college_id != ctx['college']:
            return error("无权限：仅能删除本院考试", code=403)
        exam.delete()
        return success({"examId": eid})

class ExamStudentQueryView(BaseView):
    """学生端考试信息查询：仅返回已发布(status=1)且与该生已选课程相关的考试；支持分页与学期筛选"""
    def post(self, request):
        page_no, page_size, data = get_page_params(request)
        student_id = data.get('studentId') or data.get('student_id') or request.GET.get('studentId') or request.GET.get('student_id')
        if not student_id:
            info = getattr(request, 'user_info', None)
            if info and isinstance(info, dict):
                student_id = info.get('studentId') or info.get('student_id') or info.get('studentId')
        if not student_id:
            return error("studentId 必填", code=400)
        enrolled = list(AcaEnrollment.objects.filter(student_id=student_id, status__in=['0','1']).values_list('course_id', flat=True))
        if not enrolled:
            return page_response([], 0, page_no, page_size)
        qs = AcaExam.objects.filter(course_id__in=enrolled, status='1')
        if data.get('semester'):
            qs = qs.filter(semester=data['semester'])
        if data.get('courseId'):
            qs = qs.filter(course_id=data['courseId'])
        total = qs.count()
        lst = list(qs.order_by('exam_date','start_time')[(page_no-1)*page_size: page_no*page_size].values())
        return page_response(lst, total, page_no, page_size)

    def get(self, request):
        # 兼容 GET ?studentId=&semester=&pageNo=&pageSize=
        from django.http import QueryDict
        data = {k: request.GET.get(k) for k in ['studentId','semester','courseId']}
        # 构造分页
        try:
            page_no = int(request.GET.get('pageNo') or 1)
            page_size = int(request.GET.get('pageSize') or 10)
        except:
            page_no, page_size = 1, 10
        request._body = b''
        # 复用 post 逻辑
        class _R: pass
        fake = _R()
        fake.GET = request.GET
        fake.META = request.META
        fake.headers = request.headers
        fake.user_info = getattr(request, 'user_info', None)
        # 直接走查询
        student_id = data.get('studentId')
        if not student_id:
            info = getattr(request, 'user_info', None)
            if info and isinstance(info, dict):
                student_id = info.get('studentId') or info.get('student_id')
        if not student_id:
            return error("studentId 必填", code=400)
        enrolled = list(AcaEnrollment.objects.filter(student_id=student_id, status__in=['0','1']).values_list('course_id', flat=True))
        if not enrolled:
            return page_response([], 0, page_no, page_size)
        qs = AcaExam.objects.filter(course_id__in=enrolled, status='1')
        if data.get('semester'):
            qs = qs.filter(semester=data['semester'])
        if data.get('courseId'):
            qs = qs.filter(course_id=data['courseId'])
        total = qs.count()
        lst = list(qs.order_by('exam_date','start_time')[(page_no-1)*page_size: page_no*page_size].values())
        return page_response(lst, total, page_no, page_size)

# ---- 成绩 ----
def _score_course_college(course_id):
    c = AcaCourse.objects.filter(course_id=course_id).first()
    if not c or not c.dept_id:
        return None
    if c.dept_id >= 1000:
        from system_app.models import SysDept
        d = SysDept.objects.filter(dept_id=c.dept_id).first()
        if d and d.parent_id:
            return d.parent_id
    return c.dept_id

class ScoreAddView(BaseView):
    """录入/修改成绩（含权限与状态校验）。仅保存分数，状态流转见 ScoreSubmitView。"""
    def post(self, request):
        body = self.parse_body(request)
        # GPA 简算
        try:
            score = float(body.get('score',0))
        except:
            return error("score 非法", code=400)
        if score >= 90: gpa = 4.0
        elif score >= 80: gpa = 3.0
        elif score >= 70: gpa = 2.0
        elif score >= 60: gpa = 1.0
        else: gpa = 0.0
        student_id = body.get('studentId') or body.get('student_id')
        course_id = body.get('courseId') or body.get('course_id')
        semester = body.get('semester','2024-2025-1')
        exam_type = str(body.get('examType') or body.get('exam_type') or '0')
        if not student_id or not course_id:
            return error("studentId与courseId必填", code=400)
        # 必须已选且已缴费(status=1)才能录分（防未缴费/未选课幽灵分）
        if not AcaEnrollment.objects.filter(student_id=student_id, course_id=course_id, status='1').exists():
            return error("该生未选课或未缴费，无考试资格", code=400)
        ctx = _current_user_ctx(request)
        # 非管理员按学院隔离
        if not ctx['is_admin'] and ctx['college']:
            cc = _score_course_college(course_id)
            if cc is not None and cc != ctx['college']:
                return error("无权限：仅能操作本院课程成绩", code=403)
        # 补考/重修历史保留：exam_type 0正常 1补考 2重修，is_retake标记
        is_retake = '1' if exam_type in ('1','2') else '0'
        # 权限校验（基于现有成绩状态）
        existing = AcaScore.objects.filter(student_id=student_id, course_id=course_id, semester=semester, exam_type=exam_type).first()
        if existing:
            cur = existing.status
            if cur == '3':
                if not ctx['is_admin']:
                    return error("成绩已终审，仅管理员可修改", code=403)
            elif cur == '2':
                if not ctx['is_admin']:
                    return error("成绩已提交管理员终审，无法修改", code=403)
            elif cur == '1':
                if ctx['is_teacher']:
                    return error("已提交教务确认，教师无权修改", code=403)
            # cur=='0'：教师/教务/管理员均可编辑
        new_status = existing.status if existing else '0'
        # 同(学生,课程,学期,exam_type)已存在则更新，否则创建
        if existing:
            existing.score = score
            existing.gpa_point = gpa
            existing.is_retake = is_retake
            existing.exam_id = body.get('examId') or existing.exam_id
            existing.status = new_status
            existing.save(update_fields=['score','gpa_point','is_retake','exam_id','status','update_time'])
            return success({"scoreId": existing.score_id, "gpa": float(gpa), "is_retake": existing.is_retake, "status": new_status})
        else:
            sid = body.get('scoreId') or body.get('score_id') or gen_id('SCOR')
            s = AcaScore.objects.create(score_id=sid, student_id=student_id, course_id=course_id, semester=semester, exam_type=exam_type, is_retake=is_retake, exam_id=body.get('examId'), score=score, gpa_point=gpa, status=new_status)
            return success({"scoreId": s.score_id, "gpa": float(gpa), "is_retake": is_retake, "status": new_status})

class ScoreSubmitView(BaseView):
    """成绩状态流转（逐级上报）：
       teacher_submit: 0->1（教师提交，提交后教师不可改）
       jwc_confirm: 1->2（教务确认，提交后教务不可改）
       jwc_reject:  ->0（教务退回教师）
       admin_finalize: 2->3（管理员终审）
       admin_reopen: ->0（管理员重开）
    """
    def post(self, request):
        body = self.parse_body(request)
        score_id = body.get('scoreId') or body.get('score_id')
        action = body.get('action')
        if not score_id or not action:
            return error("scoreId与action必填", code=400)
        s = AcaScore.objects.filter(score_id=score_id).first()
        if not s:
            return error("成绩记录不存在", code=404)
        ctx = _current_user_ctx(request)
        # 学院隔离
        if not ctx['is_admin'] and ctx['college']:
            cc = _score_course_college(s.course_id)
            if cc is not None and cc != ctx['college']:
                return error("无权限：仅能操作本院课程成绩", code=403)
        cur = s.status
        if action == 'teacher_submit':
            if cur != '0':
                return error("仅草稿可提交", code=400)
            if not (ctx['is_teacher'] or ctx['is_admin']):
                return error("仅教师或管理员可提交", code=403)
            s.status = '1'
        elif action == 'jwc_confirm':
            if cur != '1':
                return error("仅待教务确认记录可确认", code=400)
            if not (ctx['is_jwc'] or ctx['is_admin']):
                return error("仅教务或管理员可确认", code=403)
            s.status = '2'
        elif action == 'jwc_reject':
            if cur not in ('1', '2'):
                return error("当前状态不可退回", code=400)
            if not (ctx['is_jwc'] or ctx['is_admin']):
                return error("仅教务或管理员可退回", code=403)
            s.status = '0'
        elif action == 'admin_finalize':
            if cur != '2':
                return error("仅待管理员终审记录可终审", code=400)
            if not ctx['is_admin']:
                return error("仅管理员可终审", code=403)
            s.status = '3'
        elif action == 'admin_reopen':
            if not ctx['is_admin']:
                return error("仅管理员可重开", code=403)
            s.status = '0'
        else:
            return error("未知 action", code=400)
        s.save(update_fields=['status','update_time'])
        return success({"scoreId": s.score_id, "status": s.status})

class ScoreImportView(BaseView):
    def post(self, request):
        # 需 openpyxl，解析上传 Excel
        try:
            f = request.FILES.get('file')
            if not f:
                return error("请上传文件", code=400)
            import openpyxl
            wb = openpyxl.load_workbook(f)
            ws = wb.active
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                student_id, course_id, score, semester = row[0], row[1], row[2], (row[3] if len(row)>3 else '2024-2025-1')
                exam_type = str(row[4] if len(row)>4 and row[4] else '0')
                is_retake = '1' if exam_type!='0' else '0'
                sid = gen_id('SCOR')
                score_val = float(score or 0)
                gpa = 4.0 if score_val>=90 else 3.0 if score_val>=80 else 2.0 if score_val>=70 else 1.0 if score_val>=60 else 0.0
                # 需已选校验略，导入时仅告警
                AcaScore.objects.update_or_create(student_id=str(student_id), course_id=str(course_id), semester=str(semester), exam_type=exam_type, defaults={'score_id': sid, 'score': score_val, 'gpa_point': gpa, 'is_retake': is_retake})
                count+=1
            return success({"imported": count})
        except Exception as e:
            return error(str(e))

class ScoreQueryByPageView(BaseView):
    def post(self, request):
        page_no, page_size, data = get_page_params(request)
        qs = AcaScore.objects.all()
        college=_get_jwc_college(request)
        if college:
            allowed=_allowed_dept_ids(college)
            if allowed is not None:
                allowed_courses=list(AcaCourse.objects.filter(dept_id__in=allowed).values_list('course_id', flat=True))
                qs=qs.filter(course_id__in=allowed_courses)
        if data.get('studentId'):
            qs = qs.filter(student_id=data['studentId'])
        if data.get('courseId'):
            qs = qs.filter(course_id=data['courseId'])
        if data.get('status'):
            qs = qs.filter(status=str(data['status']))
        if data.get('examType'):
            qs = qs.filter(exam_type=str(data['examType']))
        total = qs.count()
        lst = list(qs.order_by('-create_time')[(page_no-1)*page_size: page_no*page_size].values())
        return page_response(lst, total, page_no, page_size)

class ScoreRankView(BaseView):
    def post(self, request):
        from django.db.models import Avg, Count
        data = self.parse_body(request)
        course_id = data.get('courseId')
        qs = AcaScore.objects.all()
        if course_id:
            qs = qs.filter(course_id=course_id)
        # 分数段分布
        buckets = {"90-100": qs.filter(score__gte=90).count(), "80-89": qs.filter(score__gte=80, score__lt=90).count(), "60-79": qs.filter(score__gte=60, score__lt=80).count(), "0-59": qs.filter(score__lt=60).count()}
        avg = qs.aggregate(avg=Avg('score'))['avg']
        return success({"buckets": buckets, "avg": float(avg or 0), "total": qs.count()})

# ---- 班级 ----
class ClassQueryByPageView(BaseView):
    def post(self, request):
        page_no, page_size, data = get_page_params(request)
        qs = StuClass.objects.all()
        college=_get_jwc_college(request)
        if college:
            allowed=_allowed_dept_ids(college)
            if allowed is not None:
                qs=qs.filter(dept_id__in=allowed)
        if data.get('className'):
            qs = qs.filter(class_name__icontains=data['className'])
        total = qs.count()
        lst = list(qs.order_by('head_teacher_id','class_name')[(page_no-1)*page_size: page_no*page_size].values())
        return page_response(lst, total, page_no, page_size)

class ClassSaveView(BaseView):
    def post(self, request):
        body = self.parse_body(request)
        cid = body.get('classId') or body.get('class_id')
        if cid:
            StuClass.objects.filter(class_id=cid).update(class_name=body.get('className') or body.get('class_name'), dept_id=body.get('deptId'), grade=body.get('grade'), head_teacher_id=body.get('headTeacherId'))
            return success({"classId": cid})
        c = StuClass.objects.create(class_name=body['className'], dept_id=body.get('deptId'), grade=body.get('grade'), head_teacher_id=body.get('headTeacherId'))
        return success({"classId": c.class_id})

class ClassDeleteView(BaseView):
    def post(self, request, class_id=None):
        cid = class_id or self.parse_body(request).get('classId') or self.parse_body(request).get('class_id')
        if not cid:
            return error("classId 不能为空", code=400)
        StuStudent.objects.filter(class_id=cid).update(class_id=None)
        StuClass.objects.filter(class_id=cid).delete()
        return success()

class StudentDeleteView(BaseView):
    def post(self, request, student_id=None):
        sid = student_id or self.parse_body(request).get('studentId') or self.parse_body(request).get('student_id')
        if not sid:
            return error("studentId 不能为空", code=400)
        try:
            stu = StuStudent.objects.get(student_id=sid)
        except StuStudent.DoesNotExist:
            return error("学生不存在", code=404)
        # 级联清理本库关联
        StuStudentFile.objects.filter(student_id=sid).delete()
        AcaEnrollment.objects.filter(student_id=sid).delete()
        AcaScore.objects.filter(student_id=sid).delete()
        # 清理 resource 库（宿舍分配/借阅）
        try:
            from resource_app.models import ResDormAssign, ResBorrow, ResRoom
            assigns = list(ResDormAssign.objects.using('resource').filter(student_id=sid))
            for a in assigns:
                try:
                    room = ResRoom.objects.using('resource').get(room_id=a.room_id)
                    ResRoom.objects.using('resource').filter(room_id=a.room_id).update(occupied=max(0, room.occupied-1))
                except Exception:
                    pass
            ResDormAssign.objects.using('resource').filter(student_id=sid).delete()
            ResBorrow.objects.using('resource').filter(student_id=sid).delete()
        except Exception:
            pass
        # 清理 finance 库（订单/退款）
        try:
            from finance_app.models import FeeOrder, FeeOrderItem, FeeRefund
            orders = list(FeeOrder.objects.using('finance').filter(student_id=sid))
            oids = [o.order_id for o in orders]
            FeeOrderItem.objects.using('finance').filter(order_id__in=oids).delete()
            FeeRefund.objects.using('finance').filter(order_id__in=oids).delete()
            FeeOrder.objects.using('finance').filter(student_id=sid).delete()
        except Exception:
            pass
        stu.delete()
        return success()

class StudentArchiveView(BaseView):
    """POST /student/archive {studentId} 成绩→归档，需校验（如学分/GPA）此处简化为需至少1条成绩"""
    def post(self, request):
        body = self.parse_body(request)
        sid = body.get('studentId') or body.get('student_id')
        if not sid:
            return error("studentId 不能为空", code=400)
        try:
            stu = StuStudent.objects.get(student_id=sid)
        except StuStudent.DoesNotExist:
            return error("学生不存在", code=404)
        if stu.is_final == '1':
            return error("已归档", code=400)
        if not AcaScore.objects.filter(student_id=sid).exists():
            return error("无成绩记录，无法归档", code=400)
        stu.is_final = '1'
        stu.save(update_fields=['is_final'])
        return success({"studentId": sid, "is_final": "1"})

class SchedulingDeleteView(BaseView):
    def post(self, request, id=None):
        sid = id or self.parse_body(request).get('id')
        if not sid:
            return error("id 不能为空", code=400)
        AcaScheduling.objects.filter(id=sid).delete()
        return success()

class EnrollmentUpdateView(BaseView):
    def post(self, request):
        body = self.parse_body(request)
        enroll_id = body.get('enrollId') or body.get('enroll_id')
        if not enroll_id:
            return error("enrollId 不能为空", code=400)
        try:
            en = AcaEnrollment.objects.get(enroll_id=enroll_id)
        except AcaEnrollment.DoesNotExist:
            return error("选课记录不存在", code=404)
        if 'scheduleId' in body:
            en.schedule_id = body.get('scheduleId') or None
        if 'status' in body and body['status'] is not None and body['status'] != '':
            new_status = str(body['status'])
            # 仅允许 0→2 退选，其他变迁需走缴费/退费专用接口
            allowed = {('0','2')}
            if (en.status, new_status) not in allowed:
                return error(f"非法的状态流转 {en.status}→{new_status}，仅允许 0→2", code=400)
            en.status = new_status
        en.save()
        return success({"enrollId": enroll_id})

class ScoreDeleteView(BaseView):
    def post(self, request, score_id=None):
        sid = score_id or self.parse_body(request).get('scoreId') or self.parse_body(request).get('score_id')
        if not sid:
            return error("scoreId 不能为空", code=400)
        AcaScore.objects.filter(score_id=sid).delete()
        return success()

class AttendanceMarkView(BaseView):
    """POST /attendance/mark {studentId,courseId,scheduleId} 缴费后才能上课"""
    def post(self, request):
        body = self.parse_body(request)
        sid = body.get('studentId') or body.get('student_id')
        cid = body.get('courseId') or body.get('course_id')
        sch_id = body.get('scheduleId') or body.get('schedule_id')
        if not sid or not cid:
            return error("studentId与courseId必填", code=400)
        # 必须已选且已缴费 status=1
        if not AcaEnrollment.objects.filter(student_id=sid, course_id=cid, status='1').exists():
            return error("未缴费或未选课，禁止上课（需先完成缴费 status=1）", code=400)
        # 退选后也禁止
        if AcaEnrollment.objects.filter(student_id=sid, course_id=cid, status='2').exists():
            return error("已退选，无上课资格", code=400)
        att = AcaAttendance.objects.create(student_id=sid, course_id=cid, schedule_id=sch_id, attend_status='1')
        return success({"attendanceId": att.id})

class AttendanceQueryView(BaseView):
    def post(self, request):
        page_no, page_size, data = get_page_params(request)
        qs = AcaAttendance.objects.all()
        if data.get('studentId'): qs = qs.filter(student_id=data['studentId'])
        if data.get('courseId'): qs = qs.filter(course_id=data['courseId'])
        if data.get('sessionId'): qs = qs.filter(session_id=data['sessionId'])
        if data.get('scheduleId'): qs = qs.filter(schedule_id=data['scheduleId'])
        total = qs.count()
        lst = list(qs.order_by('-create_time')[(page_no-1)*page_size: page_no*page_size].values())
        return page_response(lst, total, page_no, page_size)
# ---- 教室（教学楼） ----
class ClassroomQueryByPageView(BaseView):
    def post(self, request):
        page_no, page_size, data = get_page_params(request)
        qs = AcaClassroom.objects.all()
        if data.get('collegeId'):
            qs = qs.filter(college_id=data['collegeId'])
        if data.get('roomNo'):
            qs = qs.filter(room_no__icontains=data['roomNo'])
        total = qs.count()
        lst = list(qs.order_by('room_no')[(page_no-1)*page_size: page_no*page_size].values())
        return page_response(lst, total, page_no, page_size)

class ClassroomSaveView(BaseView):
    def post(self, request):
        body = self.parse_body(request)
        rid = body.get('classroomId') or body.get('classroom_id')
        room_no = body.get('roomNo') or body.get('room_no')
        if not room_no:
            return error("roomNo required", code=400)
        if rid:
            AcaClassroom.objects.filter(classroom_id=rid).update(college_id=body.get('collegeId'), room_no=room_no, floor=body.get('floor'), capacity=body.get('capacity'))
            return success({"classroomId": int(rid)})
        obj = AcaClassroom.objects.create(college_id=body.get('collegeId'), room_no=room_no, floor=body.get('floor'), capacity=body.get('capacity') or 50)
        return success({"classroomId": obj.classroom_id})

class ClassroomDeleteView(BaseView):
    def post(self, request, classroom_id=None):
        rid = classroom_id or self.parse_body(request).get('classroomId') or self.parse_body(request).get('classroom_id')
        if not rid:
            return error("classroomId required", code=400)
        AcaClassroom.objects.filter(classroom_id=rid).delete()
        return success()


# ---- 培养方案 ----
class PlanQueryByMajorView(BaseView):
    def post(self, request):
        body = self.parse_body(request)
        major_id = body.get('majorId') or body.get('major_id')
        qs = AcaPlan.objects.all()
        if major_id:
            qs = qs.filter(major_id=major_id)
        lst = list(qs.order_by('year_no','term').values())
        cinfo = {c.course_id: c for c in AcaCourse.objects.all()}
        out = []
        for p in lst:
            c = cinfo.get(p['course_id'])
            p['course_name'] = c.course_name if c else 'unknown'
            p['credit'] = float(c.credit) if c and c.credit is not None else (float(p['credit']) if p['credit'] else 0)
            p['course_type'] = c.course_type if c else ''
            out.append(p)
        return success(out)

class PlanSaveView(BaseView):
    def post(self, request):
        body = self.parse_body(request)
        major_id = body.get('majorId') or body.get('major_id')
        course_id = body.get('courseId') or body.get('course_id')
        if not major_id or not course_id:
            return error("majorId and courseId required", code=400)
        credit = body.get('credit')
        if credit is None:
            c = AcaCourse.objects.filter(course_id=course_id).first()
            credit = float(c.credit) if c and c.credit else 0
        obj, _ = AcaPlan.objects.update_or_create(major_id=major_id, course_id=course_id, defaults={'year_no': body.get('yearNo',1), 'term': body.get('term',1), 'is_required': body.get('isRequired','1'), 'credit': credit})
        return success({"planId": obj.plan_id})

class PlanDeleteView(BaseView):
    def post(self, request, plan_id=None):
        pid = plan_id or self.parse_body(request).get('planId') or self.parse_body(request).get('plan_id')
        if not pid:
            return error("planId required", code=400)
        AcaPlan.objects.filter(plan_id=pid).delete()
        return success()

class PlanStudentView(BaseView):
    """GET /plan/queryStudentPlan?studentId= 返回学生个人培养方案与总学分"""
    def get(self, request):
        sid = request.GET.get('studentId') or request.GET.get('student_id')
        try:
            stu = StuStudent.objects.get(student_id=sid)
        except StuStudent.DoesNotExist:
            return error("student not found", code=404)
        major_id = None
        if stu.class_id:
            c = StuClass.objects.filter(class_id=stu.class_id).first()
            if c: major_id = c.dept_id
        if not major_id:
            return error("student has no major", code=400)
        plans = list(AcaPlan.objects.filter(major_id=major_id).values())
        cinfo = {c.course_id: c for c in AcaCourse.objects.all()}
        total = 0.0
        for p in plans:
            c = cinfo.get(p['course_id'])
            p['course_name'] = c.course_name if c else 'unknown'
            p['course_type'] = c.course_type if c else ''
            cr = float(c.credit) if c and c.credit is not None else (float(p['credit']) if p['credit'] else 0)
            p['credit'] = cr
            total += cr
        return success({"majorId": major_id, "totalCredit": total, "list": plans})


# ---- 排课发布 / 按专业批量排课 ----
class SchedulingPublishView(BaseView):
    def post(self, request):
        body = self.parse_body(request)
        published = '1' if str(body.get('published', '1')) == '1' else '0'
        ids = body.get('ids') or []
        major_id = body.get('majorId')
        semester = body.get('semester')
        q = AcaScheduling.objects.all()
        if ids:
            q = q.filter(id__in=ids)
        else:
            if major_id: q = q.filter(major_id=major_id)
            if semester: q = q.filter(semester=semester)
        n = q.update(is_published=published)
        return success({"updated": n, "published": published})

class SchedulingBulkForMajorView(BaseView):
    """POST /scheduling/bulkForMajor {majorId, semester} 把该专业培养方案里的必修课批量排课"""
    def post(self, request):
        body = self.parse_body(request)
        major_id = body.get('majorId') or body.get('major_id')
        semester = body.get('semester')
        if not major_id or not semester:
            return error("majorId and semester required", code=400)
        plans = AcaPlan.objects.filter(major_id=major_id, is_required='1')
        if not plans.exists():
            return error("major has no plan", code=400)
        from system_app.models import SysUser, SysDept
        college = SysDept.objects.filter(dept_id=major_id).first()
        college_id = college.parent_id if college and college.parent_id else major_id
        teachers = list(SysUser.objects.filter(dept_id__in=_allowed_dept_ids(college_id), user_type__in=['1','8'], status='0', del_flag='0').values_list('user_id', flat=True))
        rooms = list(AcaClassroom.objects.filter(college_id=college_id))
        if not rooms:
            rooms = list(AcaClassroom.objects.all())
        if not teachers:
            teachers = list(SysUser.objects.filter(user_type__in=['1','8'], status='0', del_flag='0').values_list('user_id', flat=True))
        made = 0
        sections = ['1','2','3','4','5','6']
        for i, p in enumerate(plans):
            c = AcaCourse.objects.filter(course_id=p.course_id).first()
            if not c: continue
            teacher = teachers[i % len(teachers)] if teachers else None
            room = rooms[i % len(rooms)] if rooms else None
            weekday = (i % 5) + 1
            section = sections[i % len(sections)]
            cap = room.capacity if room else 50
            AcaScheduling.objects.update_or_create(major_id=major_id, course_id=p.course_id, semester=semester, weekday=weekday, section_type=section, defaults={'teacher_id': teacher, 'classroom_id': room.classroom_id if room else None, 'start_week': 1, 'end_week': 18, 'capacity': cap, 'is_published': '0'})
            made += 1
        return success({"made": made})


# ---- 学生选课大厅（已发布排课） ----
class SchedulingSelectableView(BaseView):
    """GET /scheduling/querySelectable?studentId=&semester=&kw=&teacherKw="""
    def get(self, request):
        from system_app.models import SysUser, SysDept
        sid = request.GET.get('studentId') or request.GET.get('student_id')
        semester = request.GET.get('semester')
        kw = request.GET.get('kw') or ''
        teacher_kw = request.GET.get('teacherKw') or ''
        if not sid:
            return error("studentId required", code=400)
        try:
            stu = StuStudent.objects.get(student_id=sid)
            major_id = None
            if stu.class_id:
                c = StuClass.objects.filter(class_id=stu.class_id).first()
                if c: major_id = c.dept_id
        except StuStudent.DoesNotExist:
            return error("student not found", code=404)
        qs = AcaScheduling.objects.filter(is_published='1')
        if semester:
            qs = qs.filter(semester=semester)
        public_types = ['public_basic','general']
        sch_list = list(qs)
        cinfo = {c.course_id: c for c in AcaCourse.objects.all()}
        allowed_ids = []
        for s in sch_list:
            c = cinfo.get(s.course_id)
            if not c: continue
            if c.course_type in public_types or (s.major_id and s.major_id == major_id):
                allowed_ids.append(s.id)
        qs = AcaScheduling.objects.filter(id__in=allowed_ids)
        enrolled = list(AcaEnrollment.objects.filter(student_id=sid, status__in=['0','1']).values_list('schedule_id', flat=True))
        dinfo = {d.dept_id: d.dept_name for d in SysDept.objects.all()}
        tinfo = {u.user_id: u.user_name for u in SysUser.objects.all()}
        rinfo = {r.classroom_id: r for r in AcaClassroom.objects.all()}
        out = []
        for s in qs:
            c = cinfo.get(s.course_id)
            if not c: continue
            if kw and kw not in (c.course_name or '') and kw not in (c.course_code or ''):
                continue
            tname = tinfo.get(s.teacher_id, '') or ''
            if teacher_kw and teacher_kw not in tname:
                continue
            if s.id in enrolled:
                continue
            room = rinfo.get(s.classroom_id)
            cnt = AcaEnrollment.objects.filter(schedule_id=s.id, status__in=['0','1']).count()
            out.append({
                'scheduleId': s.id, 'courseId': s.course_id, 'courseName': c.course_name,
                'courseCode': c.course_code, 'hours': c.hours, 'credit': float(c.credit) if c.credit else 0,
                'courseType': c.course_type, 'collegeId': c.dept_id, 'collegeName': dinfo.get(c.dept_id, ''),
                'teacherId': s.teacher_id, 'teacherName': tname,
                'classroomId': s.classroom_id, 'roomNo': room.room_no if room else '',
                'weekday': s.weekday, 'sectionType': s.section_type, 'startWeek': s.start_week, 'endWeek': s.end_week,
                'capacity': s.capacity, 'enrolled': cnt, 'remaining': (s.capacity or 0) - cnt,
            })
        return success(out)


# ---- 学生个人课表 ----
class SchedulingStudentTimetableView(BaseView):
    """GET /scheduling/queryStudentTimetable?studentId=&semester="""
    def get(self, request):
        from system_app.models import SysUser, SysDept
        sid = request.GET.get('studentId') or request.GET.get('student_id')
        semester = request.GET.get('semester')
        ens = AcaEnrollment.objects.filter(student_id=sid, status__in=['0','1'])
        sch_ids = [e.schedule_id for e in ens if e.schedule_id]
        cinfo = {c.course_id: c for c in AcaCourse.objects.all()}
        dinfo = {d.dept_id: d.dept_name for d in SysDept.objects.all()}
        tinfo = {u.user_id: u.user_name for u in SysUser.objects.all()}
        rinfo = {r.classroom_id: r for r in AcaClassroom.objects.all()}
        out = []
        for s in AcaScheduling.objects.filter(id__in=sch_ids):
            if semester and s.semester != semester:
                continue
            c = cinfo.get(s.course_id)
            room = rinfo.get(s.classroom_id)
            out.append({
                'scheduleId': s.id, 'courseId': s.course_id, 'courseName': c.course_name if c else '',
                'teacherId': s.teacher_id, 'teacherName': tinfo.get(s.teacher_id, ''),
                'roomNo': room.room_no if room else '', 'weekday': s.weekday, 'sectionType': s.section_type,
                'startWeek': s.start_week, 'endWeek': s.end_week, 'semester': s.semester,
            })
        return success(out)


# ================= 考勤签到 =================
def _gen_code(n=6):
    import random, string
    return ''.join(random.choices(string.digits, k=n))

class AttendanceSessionCreateView(BaseView):
    """教师发起签到场次：POST /attendance/session/create {scheduleId, minutes}"""
    def post(self, request):
        from datetime import datetime, timedelta
        body = self.parse_body(request)
        schedule_id = body.get('scheduleId')
        if not schedule_id:
            return error("scheduleId 必填", code=400)
        sch = AcaScheduling.objects.filter(id=schedule_id).first()
        if not sch:
            return error("排课不存在", code=404)
        minutes = int(body.get('minutes') or 5)
        now = datetime.now()
        code = _gen_code()
        s = AcaAttendanceSession.objects.create(schedule_id=schedule_id, course_id=sch.course_id,
            teacher_id=body.get('teacherId') or sch.teacher_id, session_code=code,
            start_time=now, end_time=now+timedelta(minutes=minutes), status='0')
        # 自动生成缺勤占位（未签到者）
        enrolled = AcaEnrollment.objects.filter(schedule_id=schedule_id, status__in=['0','1']).values_list('student_id', flat=True)
        for stu in enrolled:
            AcaAttendance.objects.get_or_create(student_id=stu, schedule_id=schedule_id, session_id=s.id,
                defaults={'course_id': sch.course_id, 'attend_status': '0'})
        return success({"sessionId": s.id, "code": code, "endTime": s.end_time.strftime('%Y-%m-%d %H:%M:%S')})

class AttendanceSessionCloseView(BaseView):
    def post(self, request, session_id=None):
        sid = session_id or self.parse_body(request).get('sessionId')
        AcaAttendanceSession.objects.filter(id=sid).update(status='1')
        return success({"sessionId": int(sid)})

class AttendanceSessionQueryView(BaseView):
    def post(self, request):
        page_no, page_size, data = get_page_params(request)
        qs = AcaAttendanceSession.objects.all().order_by('-id')
        if data.get('teacherId'):
            qs = qs.filter(teacher_id=data['teacherId'])
        if data.get('courseId'):
            qs = qs.filter(course_id=data['courseId'])
        total = qs.count()
        lst = list(qs[(page_no-1)*page_size: page_no*page_size].values())
        # 附带统计
        ids = [r['id'] for r in lst]
        from django.db.models import Count, Q
        stats = {a['session_id']: a for a in AcaAttendance.objects.filter(session_id__in=ids).values('session_id').annotate(
            total=Count('id'), present=Count('id', filter=Q(attend_status='1')))}
        cinfo = {c.course_id: c.course_name for c in AcaCourse.objects.all()}
        for r in lst:
            st = stats.get(r['id'], {})
            r['present'] = st.get('present', 0); r['total'] = st.get('total', 0)
            r['courseName'] = cinfo.get(r['course_id'], r['course_id'])
        return page_response(lst, total, page_no, page_size)

class AttendanceSignInView(BaseView):
    """学生扫码/输码签到：POST /attendance/signIn {studentId, code}"""
    def post(self, request):
        body = self.parse_body(request)
        student_id = body.get('studentId'); code = str(body.get('code') or '')
        if not student_id or not code:
            return error("studentId与code必填", code=400)
        from datetime import datetime
        sess = AcaAttendanceSession.objects.filter(session_code=code).order_by('-id').first()
        if not sess:
            return error("签到码无效", code=404)
        if sess.status != '0':
            return error("该签到已结束", code=400)
        now = datetime.now()
        if sess.end_time and now > sess.end_time:
            return error("已超过签到截止时间", code=400)
        # 需选修该课程
        if not AcaEnrollment.objects.filter(student_id=student_id, course_id=sess.course_id, status__in=['0','1']).exists():
            return error("您未选修该课程，无法签到", code=403)
        rec = AcaAttendance.objects.filter(student_id=student_id, session_id=sess.id).first()
        if rec and rec.attend_status == '1':
            return error("请勿重复签到", code=400)
        AcaAttendance.objects.update_or_create(student_id=student_id, session_id=sess.id,
            defaults={'course_id': sess.course_id, 'schedule_id': sess.schedule_id, 'attend_status': '1'})
        return success({"sessionId": sess.id, "courseId": sess.course_id})

class AttendanceMyStatsView(BaseView):
    """学生考勤统计：GET /attendance/myStats?studentId="""
    def get(self, request):
        student_id = request.GET.get('studentId')
        if not student_id:
            return error("studentId 必填", code=400)
        rows = list(AcaAttendance.objects.filter(student_id=student_id).values())
        cinfo = {c.course_id: c.course_name for c in AcaCourse.objects.all()}
        for r in rows: r['courseName'] = cinfo.get(r['course_id'], r['course_id'])
        return success(rows)

# ================= 补考报名 =================
class ExamSignupAddView(BaseView):
    """补考报名：校验挂科+发布补考考试，自动生成重修费订单"""
    def post(self, request):
        import time as _t
        body = self.parse_body(request)
        student_id = body.get('studentId'); exam_id = body.get('examId')
        if not student_id or not exam_id:
            return error("studentId与examId必填", code=400)
        exam = AcaExam.objects.filter(exam_id=exam_id, status='1').first()
        if not exam:
            return error("考试不存在或未发布", code=404)
        if exam.exam_type not in ('2','3'):
            return error("该考试非补考/重修场次，无需报名", code=400)
        course_id = exam.course_id
        fail = AcaScore.objects.filter(student_id=student_id, course_id=course_id).order_by('-create_time').first()
        if not fail or float(fail.score or 0) >= 60:
            return error("仅挂科课程可报名补考", code=400)
        dup = AcaExamSignup.objects.filter(student_id=student_id, exam_id=exam_id, status__in=['0','1']).exists()
        if dup:
            return error("已报名该场补考，请勿重复报名", code=400)
        signup_id = gen_id('SIGN')
        order_id = None; fee = 0
        try:
            crs = AcaCourse.objects.get(course_id=course_id)
            fee = float(crs.credit or 3) * 100
            from finance_app.models import FeeOrder, FeeOrderItem
            order_id = gen_id('ORD')
            FeeOrder.objects.create(order_id=order_id, student_id=student_id, order_amount=fee, order_status='0',
                order_type='RETAKE', semester=exam.semester or '2024-2025-1', ch_id=signup_id, detail=f"补考报名-{crs.course_name}")
            FeeOrderItem.objects.create(item_id=gen_id('ITEM'), order_id=order_id, ref_id=signup_id,
                item_name=f"补考-{crs.course_name}", item_price=fee, item_num=1, item_amount=fee)
        except Exception:
            order_id = None
        AcaExamSignup.objects.create(signup_id=signup_id, student_id=student_id, exam_id=exam_id,
            course_id=course_id, fee_order_id=order_id, status='0')
        return success({"signupId": signup_id, "orderId": order_id, "fee": fee})

class ExamSignupPayConfirmView(BaseView):
    """模拟支付成功回调：POST /examSignup/payConfirm/{signupId}"""
    def post(self, request, signup_id=None):
        su = AcaExamSignup.objects.filter(signup_id=signup_id).first()
        if not su:
            return error("报名记录不存在", code=404)
        if su.status == '1':
            return success({"signupId": signup_id, "status": '1'})
        if su.fee_order_id:
            from finance_app.models import FeeOrder
            FeeOrder.objects.filter(order_id=su.fee_order_id).update(order_status='3')
        AcaExamSignup.objects.filter(signup_id=signup_id).update(status='1')
        return success({"signupId": signup_id, "status": '1'})

class ExamSignupCancelView(BaseView):
    def post(self, request, signup_id=None):
        su = AcaExamSignup.objects.filter(signup_id=signup_id).first()
        if not su:
            return error("报名记录不存在", code=404)
        if su.status == '1':
            return error("已缴费报名不可取消", code=400)
        AcaExamSignup.objects.filter(signup_id=signup_id).update(status='2')
        if su.fee_order_id:
            from finance_app.models import FeeOrder
            FeeOrder.objects.filter(order_id=su.fee_order_id, order_status='0').update(order_status='2')
        return success({"signupId": signup_id})

class ExamSignupQueryByPageView(BaseView):
    def post(self, request):
        page_no, page_size, data = get_page_params(request)
        qs = AcaExamSignup.objects.all().order_by('-create_time')
        college = _get_jwc_college(request)
        if college:
            allowed = _allowed_dept_ids(college)
            if allowed is not None:
                ac = list(AcaCourse.objects.filter(dept_id__in=allowed).values_list('course_id', flat=True))
                qs = qs.filter(course_id__in=ac)
        if data.get('studentId'):
            qs = qs.filter(student_id=data['studentId'])
        if data.get('status'):
            qs = qs.filter(status=data['status'])
        if data.get('courseId'):
            qs = qs.filter(course_id=data['courseId'])
        total = qs.count()
        lst = list(qs[(page_no-1)*page_size: page_no*page_size].values())
        cinfo = {c.course_id: c.course_name for c in AcaCourse.objects.all()}
        einfo = {e.exam_id: e.exam_name for e in AcaExam.objects.all()}
        for r in lst:
            r['courseName'] = cinfo.get(r['course_id'], r['course_id'])
            r['examName'] = einfo.get(r['exam_id'], r['exam_id'])
        return page_response(lst, total, page_no, page_size)

# ================= 学分/挂科预警 =================
class WarningComputeView(BaseView):
    """全量重算预警：FAIL按不及格门数分级；CREDIT对照培养方案应修进度"""
    def post(self, request):
        from django.db.models import Avg
        from finance_app.models import FeeTuitionConfig
        cfg = FeeTuitionConfig.objects.order_by('-create_time').first()
        cur_sem = (cfg.semester if cfg else '2024-2025-1').split('-')[-1]
        cur_year = int((cfg.semester if cfg else '2024-2025-1').split('-')[0])
        # 课程学分映射 + 培养方案学分
        credits = {c.course_id: float(c.credit or 3) for c in AcaCourse.objects.all()}
        cnames = {c.course_id: c.course_name for c in AcaCourse.objects.all()}
        plan_by_major = {}
        for p in AcaPlan.objects.all():
            plan_by_major.setdefault(p.major_id, 0)
            plan_by_major[p.major_id] += float(p.credit or 0)
        created = 0
        AcaWarning.objects.all().delete()
        students = StuStudent.objects.filter(is_final='0')
        scores = {}
        for s in AcaScore.objects.all():
            scores.setdefault(s.student_id, []).append(s)
        for stu in students:
            ss = scores.get(stu.student_id, [])
            fails = [x for x in ss if x.exam_type == '0' and float(x.score or 0) < 60]
            n = len(fails)
            if n:
                level = 3 if n >= 5 else 2 if n >= 3 else 1
                names = ','.join([cnames.get(x.course_id, x.course_id) for x in fails][:5])
                AcaWarning.objects.create(student_id=stu.student_id, warning_type='FAIL', level=level,
                    detail=f"{n} 门课程不及格：{names}", semester=cfg.semester if cfg else '')
                created += 1
            # 学分预警：按入学年限估算应修
            major_id = None
            if stu.class_id:
                cls = StuClass.objects.filter(class_id=stu.class_id).first()
                if cls: major_id = cls.dept_id
            total_plan = plan_by_major.get(major_id)
            if total_plan and stu.enroll_year:
                years = max(cur_year - stu.enroll_year, 0) + (1 if cur_sem == '2' else 0)
                expected = round(total_plan * min(years / 4.0, 1.0), 1)
                earned = sum(credits.get(x.course_id, 0) for x in ss if x.exam_type == '0' and float(x.score or 0) >= 60)
                if earned < expected * 0.8 and years > 0:
                    lvl = 3 if earned < expected * 0.5 else 2
                    AcaWarning.objects.create(student_id=stu.student_id, warning_type='CREDIT', level=lvl,
                        detail=f"应修约 {expected} 学分，实修 {earned:.1f} 学分", semester=cfg.semester if cfg else '')
                    created += 1
        cnt_fail = AcaWarning.objects.filter(warning_type='FAIL').count()
        cnt_credit = AcaWarning.objects.filter(warning_type='CREDIT').count()
        return success({"created": created, "fail": cnt_fail, "credit": cnt_credit})

class WarningQueryByPageView(BaseView):
    def post(self, request):
        page_no, page_size, data = get_page_params(request)
        qs = AcaWarning.objects.all().order_by('-level', '-id')
        college = _get_jwc_college(request)
        if college:
            allowed_majors = _allowed_dept_ids(college)
            if allowed_majors is not None:
                class_ids = list(StuClass.objects.filter(dept_id__in=allowed_majors).values_list('class_id', flat=True))
                sids = list(StuStudent.objects.filter(class_id__in=class_ids).values_list('student_id', flat=True))
                qs = qs.filter(student_id__in=sids)
        if data.get('studentId'):
            qs = qs.filter(student_id=data['studentId'])
        if data.get('warningType'):
            qs = qs.filter(warning_type=data['warningType'])
        if data.get('handled'):
            qs = qs.filter(handled=data['handled'])
        total = qs.count()
        lst = list(qs[(page_no-1)*page_size: page_no*page_size].values())
        sinfo = {s.student_id: s.name for s in StuStudent.objects.filter(student_id__in=[r['student_id'] for r in lst])}
        for r in lst: r['studentName'] = sinfo.get(r['student_id'], '')
        return page_response(lst, total, page_no, page_size)

class WarningHandleView(BaseView):
    def post(self, request, wid=None):
        wid = wid or self.parse_body(request).get('id')
        AcaWarning.objects.filter(id=wid).update(handled='1')
        return success({"id": int(wid)})

class WarningMyView(BaseView):
    def get(self, request):
        student_id = request.GET.get('studentId')
        if not student_id:
            return error("studentId 必填", code=400)
        rows = list(AcaWarning.objects.filter(student_id=student_id).order_by('-level').values())
        return success(rows)

# ================= 学生评教 =================
class EvaluationSaveView(BaseView):
    def post(self, request):
        body = self.parse_body(request)
        student_id = body.get('studentId'); course_id = body.get('courseId')
        rating = int(body.get('rating') or 5)
        if rating < 1 or rating > 5:
            return error("rating 取值 1-5", code=400)
        if not AcaEnrollment.objects.filter(student_id=student_id, course_id=course_id, status__in=['0','1']).exists():
            return error("仅已选课程可评教", code=403)
        semester = body.get('semester') or ''
        sch = AcaScheduling.objects.filter(course_id=course_id).order_by('-id').first()
        teacher_id = sch.teacher_id if sch else None
        if not semester:
            semester = sch.semester if sch else '2024-2025-1'
        ev = AcaEvaluation.objects.filter(student_id=student_id, course_id=course_id, semester=semester).first()
        if ev:
            ev.rating = rating; ev.items = body.get('items'); ev.comment_text = body.get('comment'); 
            ev.save(update_fields=['rating','items','comment_text','update_time'])
        else:
            AcaEvaluation.objects.create(eval_id=gen_id('EVAL'), student_id=student_id, course_id=course_id,
                teacher_id=teacher_id, semester=semester, rating=rating, items=body.get('items'), comment_text=body.get('comment'))
        return success({"ok": True})

class EvaluationPendingView(BaseView):
    """待评教列表：GET /evaluation/pending?studentId= → 已缴费但未评的课程"""
    def get(self, request):
        student_id = request.GET.get('studentId')
        if not student_id:
            return error("studentId 必填", code=400)
        ens = AcaEnrollment.objects.filter(student_id=student_id, status='1')
        done = set(AcaEvaluation.objects.filter(student_id=student_id).values_list('course_id', flat=True))
        cinfo = {c.course_id: c.course_name for c in AcaCourse.objects.all()}
        out = []
        for e in ens:
            if e.course_id in done: continue
            out.append({'courseId': e.course_id, 'courseName': cinfo.get(e.course_id, e.course_id), 'enrollId': e.enroll_id})
        return success(out)

class EvaluationMyView(BaseView):
    def get(self, request):
        student_id = request.GET.get('studentId')
        if not student_id:
            return error("studentId 必填", code=400)
        rows = list(AcaEvaluation.objects.filter(student_id=student_id).order_by('-create_time').values())
        return success(rows)

class EvaluationStatsView(BaseView):
    """评教统计：POST /evaluation/queryByPage 按课程聚合"""
    def post(self, request):
        page_no, page_size, data = get_page_params(request)
        from django.db.models import Avg, Count
        qs = AcaEvaluation.objects.values('course_id').annotate(avg_rating=Avg('rating'), cnt=Count('eval_id')).order_by('-cnt')
        college = _get_jwc_college(request)
        if college:
            allowed = _allowed_dept_ids(college)
            if allowed is not None:
                ac = list(AcaCourse.objects.filter(dept_id__in=allowed).values_list('course_id', flat=True))
                qs = [q for q in qs if q['course_id'] in set(ac)]
        total = len(qs)
        rows = list(qs[(page_no-1)*page_size: page_no*page_size])
        cinfo = {c.course_id: c.course_name for c in AcaCourse.objects.all()}
        for r in rows:
            r['courseName'] = cinfo.get(r['course_id'], r['course_id'])
            r['avgRating'] = round(float(r.pop('avg_rating') or 0), 2); r.pop('avg_rating', None)
        return page_response(rows, total, page_no, page_size)

class EvaluationDetailView(BaseView):
    def get(self, request):
        course_id = request.GET.get('courseId')
        if not course_id:
            return error("courseId 必填", code=400)
        rows = list(AcaEvaluation.objects.filter(course_id=course_id).order_by('-create_time').values('eval_id','student_id','rating','comment_text','create_time'))
        sids = {s.student_id: s.name for s in StuStudent.objects.filter(student_id__in=[r['student_id'] for r in rows])}
        for r in rows:
            r['studentName'] = sids.get(r['student_id'], '匿名')
            del r['student_id']
        return success(rows)

# ================= 请假审批流 =================
class LeaveApplyView(BaseView):
    def post(self, request):
        body = self.parse_body(request)
        student_id = body.get('studentId'); leave_type = body.get('leaveType') or '事假'
        start_date = body.get('startDate'); end_date = body.get('endDate'); reason = body.get('reason')
        if not student_id or not start_date or not end_date or not reason:
            return error("studentId/startDate/endDate/reason 必填", code=400)
        stu = StuStudent.objects.filter(student_id=student_id).first()
        head = None; cls_id = None
        if stu and stu.class_id:
            cls = StuClass.objects.filter(class_id=stu.class_id).first()
            if cls: head = cls.head_teacher_id; cls_id = cls.class_id
        lid = gen_id('LEAV')
        AcaLeaveApply.objects.create(leave_id=lid, student_id=student_id, class_id=cls_id,
            head_teacher_id=head, leave_type=leave_type, start_date=start_date, end_date=end_date, reason=reason, status='0')
        return success({"leaveId": lid, "headTeacherId": head})

class LeaveCancelView(BaseView):
    def post(self, request, leave_id=None):
        lid = leave_id or self.parse_body(request).get('leaveId')
        lv = AcaLeaveApply.objects.filter(leave_id=lid).first()
        if not lv:
            return error("请假单不存在", code=404)
        if lv.status != '0':
            return error("仅待审批状态可撤回", code=400)
        AcaLeaveApply.objects.filter(leave_id=lid).update(status='3')
        return success({"leaveId": lid})

class LeaveApproveView(BaseView):
    def post(self, request):
        body = self.parse_body(request)
        lid = body.get('leaveId'); action = body.get('action'); opinion = body.get('opinion') or ''
        approver_id = body.get('approverId')
        lv = AcaLeaveApply.objects.filter(leave_id=lid).first()
        if not lv:
            return error("请假单不存在", code=404)
        if lv.status != '0':
            return error("该申请已处理", code=400)
        ctx = _current_user_ctx(request)
        is_admin = ctx['is_admin']
        is_head = str(approver_id or ctx.get('user_id')) == str(lv.head_teacher_id)
        if not (is_admin or is_head):
            return error("无权限：仅该生辅导员或管理员可审批", code=403)
        if action not in ('approve','reject'):
            return error("action 必须为 approve/reject", code=400)
        new_status = '1' if action == 'approve' else '2'
        AcaLeaveApply.objects.filter(leave_id=lid).update(status=new_status)
        AcaLeaveApproval.objects.create(leave_id=lid, approver_id=approver_id or ctx.get('user_id'),
            action='1' if action == 'approve' else '2', opinion=opinion)
        return success({"leaveId": lid, "status": new_status})

class LeaveQueryByPageView(BaseView):
    def post(self, request):
        page_no, page_size, data = get_page_params(request)
        qs = AcaLeaveApply.objects.all().order_by('-create_time')
        if data.get('studentId'):
            qs = qs.filter(student_id=data['studentId'])
        if data.get('headTeacherId'):
            qs = qs.filter(head_teacher_id=data['headTeacherId'])
        if data.get('status'):
            qs = qs.filter(status=data['status'])
        total = qs.count()
        lst = list(qs[(page_no-1)*page_size: page_no*page_size].values())
        sinfo = {s.student_id: (s.name, s.class_id) for s in StuStudent.objects.filter(student_id__in=[r['student_id'] for r in lst])}
        tids = set()
        for r in lst:
            si = sinfo.get(r['student_id']); 
            r['studentName'] = si[0] if si else ''
            if r['head_teacher_id']: tids.add(r['head_teacher_id'])
        from system_app.models import SysUser
        tinfo = {u.user_id: u.user_name for u in SysUser.objects.filter(user_id__in=tids)}
        for r in lst: r['headTeacherName'] = tinfo.get(r['head_teacher_id'], '')
        return page_response(lst, total, page_no, page_size)


# ================= Excel 导出 =================
def _xlsx_response(rows, headers, filename):
    from openpyxl import Workbook
    from io import BytesIO
    from django.http import HttpResponse
    wb = Workbook(); ws = wb.active; ws.title = 'sheet1'
    ws.append(headers)
    for r in rows: ws.append([r.get(h) if isinstance(r, dict) else r for h in headers])
    buf = BytesIO(); wb.save(buf)
    from urllib.parse import quote
    resp = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}.xlsx"
    return resp

class ScoreExportView(BaseView):
    """成绩单导出：GET /score/export?courseId=&studentId="""
    def get(self, request):
        qs = AcaScore.objects.all().order_by('-create_time')
        college = _get_jwc_college(request)
        if college:
            allowed = _allowed_dept_ids(college)
            if allowed is not None:
                ac = list(AcaCourse.objects.filter(dept_id__in=allowed).values_list('course_id', flat=True))
                qs = qs.filter(course_id__in=ac)
        cid = request.GET.get('courseId'); sid = request.GET.get('studentId')
        if cid: qs = qs.filter(course_id=cid)
        if sid: qs = qs.filter(student_id=sid)
        cinfo = {c.course_id: c.course_name for c in AcaCourse.objects.all()}
        sinfo = {s.student_id: s.name for s in StuStudent.objects.filter(student_id__in=set(qs.values_list('student_id', flat=True)[:5000]))}
        rows = []
        for s in qs[:5000]:
            rows.append({'学号': s.student_id, '姓名': sinfo.get(s.student_id,''), '课程': cinfo.get(s.course_id,s.course_id),
                '分数': float(s.score or 0), '绩点': float(s.gpa_point or 0), '学期': s.semester,
                '类型': {'0':'正常','1':'补考','2':'重修'}.get(s.exam_type, s.exam_type),
                '状态': {'0':'草稿','1':'待教务确认','2':'待终审','3':'已终审'}.get(s.status, s.status)})
        return _xlsx_response(rows, ['学号','姓名','课程','分数','绩点','学期','类型','状态'], '成绩单')

class EnrollmentRosterExportView(BaseView):
    """选课名单导出：GET /enrollment/export?courseId="""
    def get(self, request):
        qs = AcaEnrollment.objects.all().order_by('-create_time')
        college = _get_jwc_college(request)
        if college:
            allowed = _allowed_dept_ids(college)
            if allowed is not None:
                ac = list(AcaCourse.objects.filter(dept_id__in=allowed).values_list('course_id', flat=True))
                qs = qs.filter(course_id__in=ac)
        cid = request.GET.get('courseId'); status = request.GET.get('status')
        if cid: qs = qs.filter(course_id=cid)
        if status: qs = qs.filter(status=status)
        cinfo = {c.course_id: c.course_name for c in AcaCourse.objects.all()}
        students = StuStudent.objects.filter(student_id__in=set(qs.values_list('student_id', flat=True)[:5000]))
        cls_ids = set(students.values_list('class_id', flat=True))
        clinfo = {c.class_id: c.class_name for c in StuClass.objects.filter(class_id__in=cls_ids)}
        sinfo = {s.student_id: (s.name, clinfo.get(s.class_id,'')) for s in students}
        rows = []
        for e in qs[:5000]:
            si = sinfo.get(e.student_id, ('',''))
            rows.append({'学号': e.student_id, '姓名': si[0], '班级': si[1], '课程': cinfo.get(e.course_id,e.course_id),
                '状态': {'0':'待缴费','1':'已缴费','2':'已退选'}.get(e.status, e.status), '时间': str(e.create_time)[:19]})
        return _xlsx_response(rows, ['学号','姓名','班级','课程','状态','时间'], '选课名单')
